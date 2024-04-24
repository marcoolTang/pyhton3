# -*- coding: utf-8 -*-
from flask import Flask, render_template, send_file, jsonify, request, session
import requests
import openpyxl
from openpyxl.styles import Alignment
import os
import logging
import time
import uuid
from threading import Thread

app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'

# 获取当前脚本所在目录的绝对路径
script_dir = os.path.dirname(os.path.abspath(__file__))

# 设置日志文件路径为相对于当前脚本目录的路径
log_path = os.path.join(script_dir, "app.log")

logging.basicConfig(filename=log_path, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 定义一个字典来存储每个用户的搜索状态和 Excel 文件路径
search_status = {}

def fetch_data(query, page=1, size=100):
    url = f"https://b2b.baidu.com/s/a?q={query}&ajax=1&p={page}&s={size}"
    response = requests.get(url)
    data = response.json()
    logging.info(f"Fetched data for query '{query}', page {page}: {data}")
    return data.get('data', {}).get('productList', []), data.get('data', {}).get('dispNum', 0)

def generate_excel(data, query):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product List"
    ws['A1'] = "产品名"
    ws['B1'] = "价格"
    ws['C1'] = "跳转地址"
    ws['D1'] = "商家名"
    ws['E1'] = "来源页面地址"

    # 设置列宽和自动换行
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    for idx, product in enumerate(data, start=2):
        price = product.get('price', '')
        currency = product.get('pCurrency', '')
        unit = product.get('unit', '')
        ws[f'A{idx}'] = query
        ws[f'B{idx}'] = f"{price} {currency}/{unit}"
        ws[f'C{idx}'] =  product.get('jUrl', '')
        ws[f'D{idx}'] = product.get('fullProviderName', '')
        ws[f'E{idx}'] = f"https://b2b.baidu.com/ 爱采购"
    return wb

def fetch_all_data(query, size=60):
    page = 1
    all_data = []
    total_data_count = 0
    while True:
        data, dispNum = fetch_data(query, page, size)
        total_data_count += len(data)
        all_data.extend(data)
        print(f"Page {page}: Fetched {len(data)} items, total: {total_data_count}")
        if len(data) >= dispNum:
            break
        page += 1
        time.sleep(1)  # 延迟1秒，避免对服务器造成过大的负载
    return all_data

def scrape_and_generate_excel(query, session_id):
    # 生成文件名
    filename = f"{session_id}.xlsx"
    excel_dir = os.path.join(app.root_path, 'static')
    excel_path = os.path.join(excel_dir, filename)

    all_data = fetch_all_data(query)

    # 生成 Excel 文件
    wb = generate_excel(all_data, query)

    # 保存 Excel 文件
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    wb.save(excel_path)

    # 更新搜索状态字典
    search_status[session_id] = {'completed': True, 'excel_link': f"/download/{filename}"}

@app.route('/')
def index():
    return render_template('index2.html')

@app.route('/start_scrape')
def start_scrape():
    query = request.args.get('query', '猪肉')
    session_id = str(uuid.uuid4())
    session['session_id'] = session_id
    # 启动后台线程执行搜索和生成 Excel 文件
    thread = Thread(target=scrape_and_generate_excel, args=(query, session_id))
    thread.start()
    return jsonify({'message': 'Scraping started successfully', 'status': 'pending'})

@app.route('/check_status')
def check_status():
    session_id = session.get('session_id', None)
    if session_id in search_status:
        return jsonify(search_status[session_id])
    else:
        return jsonify({'completed': False, 'message': 'Session ID not found'})

@app.route('/download/<filename>')
def download(filename):
    return send_file(os.path.join(app.root_path, 'static', filename), as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
