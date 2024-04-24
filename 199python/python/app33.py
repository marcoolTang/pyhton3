from flask import Flask, render_template, jsonify, send_file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import openpyxl
import threading
import os
import logging
import mysql.connector
from flask_cors import CORS

app = Flask(__name__)

# 配置日志记录
logging.basicConfig(filename='app.log', level=logging.ERROR)

# 爬取任务的状态
scraping_status = {
    'status': 'idle',
    'data': [],
    'download_link': None
}

def scrape_data():
    try:
        global scraping_status
        # 创建一个 Chrome 浏览器实例

        # 设置 Chrome 选项
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # 启用无头模式
        chrome_options.add_argument('--disable-gpu')  # 禁用 GPU 加速
        chrome_options.add_argument('--no-sandbox')  # 以沙盒模式运行

        # 创建 Chrome WebDriver
        driver = webdriver.Chrome(options=chrome_options)

        # 打开页面
        url = "https://cn.vuejs.org/"  # 你要爬取的页面URL
        driver.get(url)

        # 等待页面加载完成
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "accent"))
        )

        # 获取页面信息
        price_elements = driver.find_elements(By.CLASS_NAME, "accent")
        title_elements = driver.find_elements(By.CLASS_NAME, "description")

        data = []
        for  title, price in zip(title_elements, price_elements):
            data.append({'title': title.text, 'price': price.text})

        # 关闭浏览器
        driver.quit()

        # 获取当前脚本所在目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        excel_dir = os.path.join(current_dir, 'static')

        # 确保 static 目录存在
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

    
        # 创建一个 Excel 文件并写入数据
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Page Info"
        ws['A1'] = "Title"
        ws['B1'] = "Price"
        for idx, item in enumerate(data, start=2):
            ws[f'A{idx}'] = item['title']
            ws[f'B{idx}'] = item['price']

        # Excel 文件路径
        excel_filename = os.path.join(excel_dir, "page_info.xlsx")

        wb.save(excel_filename)

        # 更新爬取任务的状态
        scraping_status['status'] = 'done'
        scraping_status['data'] = data
        scraping_status['download_link'] = f"/download/{excel_filename}"
    except Exception as e:
        # 记录异常到日志
        logging.error(f"Scraping error: {str(e)}")
        # 更新爬取任务的状态
        scraping_status['status'] = 'error'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start_scrape')
def start_scrape():
    global scraping_status
    if scraping_status['status'] == 'idle':
        scraping_status['status'] = 'scraping'
        thread = threading.Thread(target=scrape_data)
        thread.start()
    return 'OK'

@app.route('/check_status')
def check_status():
    global scraping_status
    return jsonify(scraping_status)

@app.route('/download/<path:filename>')
def download(filename):
    return send_file(os.path.join('static', filename), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
